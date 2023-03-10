from pathlib import Path
from types import NoneType
from openpyxl import Workbook
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from rich.prompt import Prompt
from rich.console import Console

from console_utils import print_in_table
from model import CourseWork
from utils import is_number


error_console = Console(stderr=True, style="bold red")


def open_file() -> Workbook:
    workbook: Workbook | None = None
    while workbook == None:
        try:
            file_path = Prompt.ask(
                "RAW Mark-sheet", default="C:\\Users\\Temp\\Desktop\\test.xlsx")
            file_path = file_path.strip('\"')
            if Path(file_path).is_file():
                workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            error_console.print("Error:", e)
    return workbook


def read_numeric_column(sheet: Worksheet, col: str) -> list[str]:
    """
    Note that this function will also convert the numeric values to string
    """
    result = []
    raw = list([it.value for it in sheet[col]])
    for it in raw:
        if is_number(it):
            result.append(str(it))
    return result


def find_marks_column(sheet: Worksheet, course_work: str):
    for col in sheet.iter_cols():
        for cell in col:
            if type(cell.value) is str and cell.value.lower() == course_work.lower():
                return cell.column_letter


def find_student_column(sheet: Worksheet):
    for col in sheet.iter_cols():
        for cell in col:
            if type(cell.value) is str and cell.value.lower().startswith('student no'):
                return cell.column_letter
